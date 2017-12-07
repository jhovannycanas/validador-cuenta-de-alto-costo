#!/usr/bin/env python
# -*- coding: utf-8 -*-

import pandas as pd
from validacion import metodos, valores
from validacion import constansts
import numpy as np
import collections
from openpyxl.styles import PatternFill
import datetime

class filevalidation():
    def __init__(self, file, tipo):
        if tipo == constansts.ARCHIVO_EXCEL:
            self.tipo = tipo
            #dtype=object,
            self.data = pd.read_excel(file,dtype=object)
        if tipo == constansts.ARCHIVO_CSV:
            self.tipo = tipo
            self.data = pd.read_csv(file,header = None,dtype=object ,sep='\t')
            self.data.columns = valores.dict_columnasMalla.values()
        self._columnas = self.data.columns
        self.dfvalidacion = pd.DataFrame(columns=self._columnas)
        #self.data = self.data.apply(lambda x: x.str.strip() if x.dtype == "object" else x)
        self.errores = []


    # def ajustarfechastimestamp(self):
    #     for fecha in metodos.fechas_convertir:
    #             self.data[self.data.columns[fecha]] = self.data[self.data.columns[fecha]].apply(
    #                 lambda x: metodos.ajustarfecha(x))

    def ajustarfechastimestamp(self):
        for fecha in metodos.fechas_convertir:
                self.data[self.data.columns[fecha]] = self.data[self.data.columns[fecha]].apply(
                    lambda x: self.ajustarfecha(x))

    def ajustarfecha(self,fecha):
        if isinstance(fecha, pd.Timestamp):
            if fecha.year <= 1900:
                fecha = fecha - datetime.timedelta(days=1)
                return fecha
            else:
                return fecha
        else:
            return fecha

    def crearexcelerrores(self, archivo):
        self.writer = pd.ExcelWriter(archivo, engine='openpyxl')
        self.data.to_excel(self.writer,index=None, sheet_name="errores",header=True)
        self.workbook = self.writer.book
        self.worksheet = self.writer.sheets['errores']
        for indice in range(0,len(self.data.index)):
            for columna in range(0, len(self.data.columns)):
                if self.dfvalidacion.values[indice, columna] == False:
                    self.worksheet.cell(row=indice+2,column=columna+1).fill = PatternFill(start_color= 'ffff00', end_color= 'ffff00', fill_type="solid")
                    #self.worksheet.write(indice, columna, self.data.iloc[indice, columna], formato1)

        self.writer.save()


    def validarunicovalor(self):
        return len(self.data[self.data.columns[0]].value_counts())>1

    def convertirafechas(self):
        for fecha in metodos.fechas_convertir:
            self.data[self.data.columns[fecha]] = pd.to_datetime(self.data[self.data.columns[fecha]], yearfirst=True,format='%Y-%m-%d').dt.date

    def convertirnumero(self):
        for numero in metodos.numeros_convertir:
            self.data[self.data.columns[numero]] = pd.to_numeric(self.data[self.data.columns[numero]])

    def validarLongitudColumnas(self):
        return len(self.data.columns) == constansts.NUMERO_COLUMNAS

    def validarCamposNulos(self):
        if self.data.isnull().any().any():
            self.nulos = self.data.notnull()
            for columna in self.data.columns:
                self.crearInformeErrores(self.nulos, self.data,
                                         columna, constansts.ERROR_CAMPOS_NULOS, self.tipo)
            return False
        return True

    def validarSecuenciaColumnas(self):
        ordenColumnas = metodos.validarOrdenColumnas(self.data, valores.dict_columnasMalla)
        if len(ordenColumnas) > 0:
            for columnaDesOrdenada in ordenColumnas:
                self.errores.append(metodos.dict_mensajes.get(constansts.ERROR_COLUMNAS_ORDEN) % (
                self.dfvalidacion.columns[columnaDesOrdenada].encode('utf-8'),
                valores.dict_columnasMalla.get(columnaDesOrdenada)))
                return False

        else:
            return True

    def crearInformeErrores(self, dfValidacion, dfData, columna, mensaje, tipo):
        if metodos.validarInconsistenciaColumna(dfValidacion, columna):
            self.errores.append(metodos.generarMensajeInconsistencia(
                dfValidacion, dfData, columna, mensaje, tipo))

    def validarFiltroTexto(self):
        # con este se puede enviar losparametros para identificar la posicion que se
        # preseta la inconsistenciahttps://stackoverflow.com/questions/12182744/python-pandas-apply-a-function-with-arguments-to-a-series
        #dg = self.data[self.data.columns[0]].as_matrix()
        #print(len(dg))
        self.validarCaracteresEspeciales()
        return False in self.dfvalidacion.values
        #return False in self.dfvalidacion.loc[:, self._columnas[12]].values

    def validarCaracteresEspeciales(self):
        self.dfvalidacion = self.data.applymap(lambda x: metodos.verificarSoloCaracteres(x))
        for columna in self.dfvalidacion.columns:
            # if self._columnas[12] == columna:
            #     self.dfvalidacion[self.dfvalidacion.columns[12]] = True
            #     continue
            self.crearInformeErrores(self.dfvalidacion, self.data, columna,
                                     constansts.ERROR_CARACTERES_ESPECIALES, self.tipo)

    def removerespacios_convertmayuscula(self):
        for columna in metodos.mayusculas_noespacio:
                self.data[self.data.columns[columna]] = self.data[self.data.columns[columna]].str.strip()
                self.data[self.data.columns[columna]] = self.data[self.data.columns[columna]].str.upper()

    def obtenerEAPB(self):
        return self.data[self.dfvalidacion.columns[0]].iloc[0]


#se cambia dfvalidacion to self.data
    def validarfechas(self):
        for fecha in metodos.fechas_convertir:
            self.dfvalidacion[self.dfvalidacion.columns[fecha]] = self.data[self.data.columns[fecha]].apply(
            lambda x: metodos.validarFormatoFecha(x))
            self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[fecha],
                                     constansts.ERROR_FORMATO_FECHA, self.tipo)

    def validarnumeros(self):
        for numero in metodos.numeros_convertir:
            self.dfvalidacion[self.dfvalidacion.columns[numero]] = self.data[self.data.columns[numero]].apply(
            lambda x: metodos.validarSoloNumeros(x))

            self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[numero],
                                     constansts.ERROR_SOLONUMEROS, self.tipo)


    def validar_EPS(self):
        self.dfvalidacion[self.dfvalidacion.columns[0]] = self.data[self.data.columns[0]].apply(
            lambda x: metodos.validarValorRespuesta(x, valores.dict_EAPB))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[0],
                                 constansts.ERROR_VALORNOPERMITIDO, self.tipo)

    def validar_Regimen(self):
        self.dfvalidacion[self.dfvalidacion.columns[1]] = self.data[self.data.columns[1]].apply(
            lambda x: metodos.validarValorRespuesta(x, valores.dict_Regimen))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[1],
                                 constansts.ERROR_VALORNOPERMITIDO, self.tipo)

    def validar_grupopobalcional(self):
        self.dfvalidacion[self.dfvalidacion.columns[2]] = self.data[self.data.columns[2]].apply(
            lambda x: metodos.validarValorRespuesta(x, valores.dict_grupoPoblacional))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[2],
                                 constansts.ERROR_VALORNOPERMITIDO, self.tipo)

    def validar_PrimerNombre(self):
        self.dfvalidacion[self.dfvalidacion.columns[3]] = self.data[self.data.columns[3]].apply(
            lambda x: metodos.validarSegundoNombre(x))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[3],
                                 constansts.ERROR_SEGUNDO_NOMBRE, self.tipo)

    def validar_SegundoNombre(self):
        self.dfvalidacion[self.dfvalidacion.columns[4]] = self.data[self.data.columns[4]].apply(
            lambda x: metodos.validarSegundoNombre(x))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[4],
                                 constansts.ERROR_SEGUNDO_NOMBRE, self.tipo)

    def validar_PrimerApellido(self):
        self.dfvalidacion[self.dfvalidacion.columns[5]] = self.data[self.data.columns[5]].apply(
            lambda x: metodos.validarPrimerNombre(x))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[5],
                                 constansts.ERROR_SEGUNDO_NOMBRE, self.tipo)

    def validar_SegundoApellido(self):
        self.dfvalidacion[self.dfvalidacion.columns[6]] = self.data[self.data.columns[6]].apply(
            lambda x: metodos.validarSegundoApellido(x))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[6],
                                 constansts.ERROR_SEGUNDO_NOMBRE, self.tipo)

    def validar_tipoidentificacion(self):
        self.dfvalidacion[self.dfvalidacion.columns[7]] = self.data[self.data.columns[7]].apply(
            lambda x: metodos.validarValorRespuesta(x, valores.dict_TipoIdentificacion))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[7],
                                 constansts.ERROR_VALORNOPERMITIDO, self.tipo)

    def validar_Identificacion(self):
        self.dfvalidacion[self.dfvalidacion.columns[8]] = self.data[self.data.columns[8]].apply(
            lambda x: metodos.validarSoloNumeros(x))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[8],
                                 constansts.ERROR_SOLONUMEROS, self.tipo)

    def validar_FechaNacimiento(self):
        self.dfvalidacion[self.dfvalidacion.columns[9]] = self.data[self.data.columns[9]].apply(
            lambda x: metodos.validarFormatoFecha(x))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[9],
                                 constansts.ERROR_FORMATO_FECHA, self.tipo)

    def validar_sexo(self):
        self.dfvalidacion[self.dfvalidacion.columns[10]] = self.data[self.data.columns[10]].apply(
            lambda x: metodos.validarValorRespuesta(x, valores.dict_Sexo))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[10],
                                 constansts.ERROR_VALORNOPERMITIDO, self.tipo)

    def validar_etnia(self):
        self.dfvalidacion[self.dfvalidacion.columns[11]] = self.data[self.data.columns[11]].apply(
            lambda x: metodos.validarValorRespuesta(x, valores.dict_Etnia))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[11],
                                 constansts.ERROR_VALORNOPERMITIDO, self.tipo)

    def validar_DireccionResidencia(self):
        pass

    def validar_Telefono(self):
        self.dfvalidacion[self.dfvalidacion.columns[13]] = self.data[self.data.columns[13]].apply(
            lambda x: metodos.validarSoloNumeros(x))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[13],
                                 constansts.ERROR_SOLONUMEROS, self.tipo)

    def validar_CodigoMunicipio(self):
        self.dfvalidacion[self.dfvalidacion.columns[14]] = self.data[self.data.columns[14]].apply(
            lambda x: metodos.validarValorRespuesta(str(x), valores.dict_mucnicipio_dos))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[14],
                                 constansts.ERROR_VALORNOPERMITIDO, self.tipo)

    def validar_FechaAfiliacionEAPB(self):
        self.dfvalidacion[self.dfvalidacion.columns[15]] = self.data[self.data.columns[15]].apply(
            lambda x: metodos.validarFormatoFecha(x))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[15],
                                 constansts.ERROR_FORMATO_FECHA,self.tipo)

    def validar_FechaInicioSintomas(self):
        self.dfvalidacion[self.dfvalidacion.columns[16]] = self.data[self.data.columns[16]].apply(
            lambda x: metodos.validarFechaSinDato(x))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[16],
                                 constansts.ERROR_FECHA_SINDATO, self.tipo)

    def validar_FechaPrimeraVisitaAr(self):
        self.dfvalidacion[self.dfvalidacion.columns[17]] = self.data[self.data.columns[17]].apply(
            lambda x: metodos.validarFechaSinDato(x))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[17],
                                 constansts.ERROR_FECHA_SINDATO, self.tipo)

    def validar_FechaDiagnosticaAr(self):
        self.dfvalidacion[self.dfvalidacion.columns[18]] = self.data[self.data.columns[18]].apply(
            lambda x: metodos.validarFechaSinDato(x))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[18],
                                 constansts.ERROR_FECHA_SINDATO, self.tipo)

    def validar_Talla(self):
        self.dfvalidacion[self.dfvalidacion.columns[19]] = self.data[self.data.columns[19]].apply(
            lambda x: metodos.validarRango(50, 250, x))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[19],
                                 constansts.ERROR_VALORNOPERMITIDO, self.tipo)

    def validar_PesoInicial(self):
        self.dfvalidacion[self.dfvalidacion.columns[20]] = self.data[self.data.columns[20]].apply(
            lambda x: metodos.validarRango(3, 300, x))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[20],
                                 constansts.ERROR_VALORNOPERMITIDO, self.tipo)

    def validar_Radiografiamanos(self):
        self.dfvalidacion[self.dfvalidacion.columns[21]] = self.data[self.data.columns[21]].apply(
            lambda x: metodos.validarValorRespuesta(x, valores.dict_Radiografia_Manos))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[21],
                                 constansts.ERROR_VALORNOPERMITIDO, self.tipo)

    def validar_Radiografia_Pies(self):
        self.dfvalidacion[self.dfvalidacion.columns[22]] = self.data[self.data.columns[22]].apply(
            lambda x: metodos.validarValorRespuesta(x, valores.dict_Radiografia_Pies))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[22],
                                 constansts.ERROR_VALORNOPERMITIDO, self.tipo)

    def validar_VSGInicial(self):
        self.dfvalidacion[self.dfvalidacion.columns[23]] = self.data[self.data.columns[23]].apply(
            lambda x: metodos.validarRangoSinDato(0, 250, 300, x))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[23],
                                 constansts.ERROR_VALORNOPERMITIDO, self.tipo)

    def validar_PCRInicial(self):
        self.dfvalidacion[self.dfvalidacion.columns[24]] = self.data[self.data.columns[24]].apply(
            lambda x: metodos.validarRangoSinDato(0, 250, 300, x))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[23],
                                 constansts.ERROR_VALORNOPERMITIDO, self.tipo)

    def validar_Factor_Reumatoideo_Inicial(self):
        self.dfvalidacion[self.dfvalidacion.columns[25]] = self.data[self.data.columns[25]].apply(
            lambda x: metodos.validarValorRespuesta(x, valores.dict_Factor_Reumatoideo_Inicial))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[25],
                                 constansts.ERROR_VALORNOPERMITIDO, self.tipo)

    def validar_HemglobinaInicial(self):
        self.dfvalidacion[self.dfvalidacion.columns[26]] = self.data[self.data.columns[26]].apply(
            lambda x: metodos.validarRangoSinDato(3, 50, 300, x))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[26],
                                 constansts.ERROR_VALORNOPERMITIDO, self.tipo)

    def validar_Leucocitos(self):
        self.dfvalidacion[self.dfvalidacion.columns[27]] = self.data[self.data.columns[27]].apply(
            lambda x: metodos.validarRangoSinDato(0, 20000, 22222, x))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[27],
                                 constansts.ERROR_VALORNOPERMITIDO, self.tipo)

    def validar_Creatina(self):
        self.dfvalidacion[self.dfvalidacion.columns[28]] = self.data[self.data.columns[28]].apply(
            lambda x: metodos.validarRangoSinDato(0, 20, 300, x))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[28],
                                 constansts.ERROR_VALORNOPERMITIDO, self.tipo)

    def validar_TFGInicial(self):
        self.dfvalidacion[self.dfvalidacion.columns[29]] = self.data[self.data.columns[29]].apply(
            lambda x: metodos.validarRangoSinDato(0, 250, 300, x))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[29],
                                 constansts.ERROR_VALORNOPERMITIDO, self.tipo)

    def validar_Parcial_Orina_Inicial(self):
        self.dfvalidacion[self.dfvalidacion.columns[30]] = self.data[self.data.columns[30]].apply(
            lambda x: metodos.validarValorRespuesta(x, valores.dict_Parcial_Orina_Inicial))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[30],
                                 constansts.ERROR_VALORNOPERMITIDO, self.tipo)

    def validar_ALT_Inicial(self):
        self.dfvalidacion[self.dfvalidacion.columns[31]] = self.data[self.data.columns[31]].apply(
            lambda x: metodos.validarValorRespuesta(x, valores.dict_ALT_Inicial))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[31],
                                 constansts.ERROR_VALORNOPERMITIDO, self.tipo)

    def validar_Anti_CCP(self):
        self.dfvalidacion[self.dfvalidacion.columns[32]] = self.data[self.data.columns[32]].apply(
            lambda x: metodos.validarValorRespuesta(x, valores.dict_Anti_CCP))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[32],
                                 constansts.ERROR_VALORNOPERMITIDO, self.tipo)

    def validar_HTA(self):
        self.dfvalidacion[self.dfvalidacion.columns[33]] = self.data[self.data.columns[33]].apply(
            lambda x: metodos.validarValorRespuesta(x, valores.dict_HTA))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[33],
                                 constansts.ERROR_VALORNOPERMITIDO, self.tipo)

    def validar_DM(self):
        self.dfvalidacion[self.dfvalidacion.columns[34]] = self.data[self.data.columns[34]].apply(
            lambda x: metodos.validarValorRespuesta(x, valores.dict_DM))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[34],
                                 constansts.ERROR_VALORNOPERMITIDO, self.tipo)

    def validar_ECV(self):
        self.dfvalidacion[self.dfvalidacion.columns[35]] = self.data[self.data.columns[35]].apply(
            lambda x: metodos.validarValorRespuesta(x, valores.dict_ECV))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[35],
                                 constansts.ERROR_VALORNOPERMITIDO, self.tipo)

    def validar_ERC(self):
        self.dfvalidacion[self.dfvalidacion.columns[36]] = self.data[self.data.columns[36]].apply(
            lambda x: metodos.validarValorRespuesta(x, valores.dict_ERC))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[36],
                                 constansts.ERROR_VALORNOPERMITIDO, self.tipo)

    def validar_Osteoporosis(self):
        self.dfvalidacion[self.dfvalidacion.columns[37]] = self.data[self.data.columns[37]].apply(
            lambda x: metodos.validarValorRespuesta(x, valores.dict_Osteoporosis))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[37],
                                 constansts.ERROR_VALORNOPERMITIDO, self.tipo)

    def validar_Sindrome_Sjogren(self):
        self.dfvalidacion[self.dfvalidacion.columns[38]] = self.data[self.data.columns[38]].apply(
            lambda x: metodos.validarValorRespuesta(x, valores.dict_Sindrome_Sjogren))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[38],
                                 constansts.ERROR_VALORNOPERMITIDO, self.tipo)

    def validar_FechaPrimerDAS(self):
        self.dfvalidacion[self.dfvalidacion.columns[39]] = self.data[self.data.columns[39]].apply(
            lambda x: metodos.validarFechaSinDato(x))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[39],
                                 constansts.ERROR_FORMATO_FECHA, self.tipo)

    def validar_Profesional_DAS(self):
        self.dfvalidacion[self.dfvalidacion.columns[40]] = self.data[self.data.columns[40]].apply(
            lambda x: metodos.validarValorRespuesta(x, valores.dict_Profesional_DAS))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[40],
                                 constansts.ERROR_VALORNOPERMITIDO, self.tipo)

    def validar_ResultadoPrimerDas(self):
        self.dfvalidacion[self.dfvalidacion.columns[41]] = self.data[self.data.columns[41]].apply(
            lambda x: metodos.validarRangoSinDato(0, 10, 300, x))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[41],
                                 constansts.ERROR_VALORNOPERMITIDO, self.tipo)

    def validar_FechaHAQ(self):
        self.dfvalidacion[self.dfvalidacion.columns[42]] = self.data[self.data.columns[42]].apply(
            lambda x: metodos.validarFechaSinDato(x))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[42],
                                 constansts.ERROR_FECHA_SINDATO, self.tipo)

    def validar_HAQ(self):
        self.dfvalidacion[self.dfvalidacion.columns[43]] = self.data[self.data.columns[43]].apply(
            lambda x: metodos.validarRangoSinDato(0, 3, 300, x))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[43],
                                 constansts.ERROR_VALORNOPERMITIDO, self.tipo)

    def validar_FechaDMARD(self):
        self.dfvalidacion[self.dfvalidacion.columns[44]] = self.data[self.data.columns[44]].apply(
            lambda x: metodos.validarFechaSinDato(x))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[44],
                                 constansts.ERROR_FECHA_SINDATO, self.tipo)

    def validar_Analgesicos_No_Opioides(self):
        self.dfvalidacion[self.dfvalidacion.columns[45]] = self.data[self.data.columns[45]].apply(
            lambda x: metodos.validarValorRespuesta(x, valores.dict_Analgesicos_No_Opioides))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[45],
                                 constansts.ERROR_VALORNOPERMITIDO, self.tipo)

    def validar_Analgesicos_Opioides(self):
        self.dfvalidacion[self.dfvalidacion.columns[46]] = self.data[self.data.columns[46]].apply(
            lambda x: metodos.validarValorRespuesta(x, valores.dict_Analgesicos_Opioides))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[46],
                                 constansts.ERROR_VALORNOPERMITIDO, self.tipo)

    def validar_AINES(self):
        self.dfvalidacion[self.dfvalidacion.columns[47]] = self.data[self.data.columns[47]].apply(
            lambda x: metodos.validarValorRespuesta(x, valores.dict_AINES))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[47],
                                 constansts.ERROR_VALORNOPERMITIDO, self.tipo)

    def validar_Corticoides(self):
        self.dfvalidacion[self.dfvalidacion.columns[48]] = self.data[self.data.columns[48]].apply(
            lambda x: metodos.validarValorRespuesta(x, valores.dict_Corticoides))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[48],
                                 constansts.ERROR_VALORNOPERMITIDO, self.tipo)

    def validar_FechaTratamientoDMARD(self):
        self.dfvalidacion[self.dfvalidacion.columns[49]] = self.data[self.data.columns[49]].apply(
            lambda x: metodos.validarFechaSinDato(x))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[49],
                                 constansts.ERROR_FECHA_SINDATO, self.tipo)

    def validar_Tamizaje_TB(self):
        self.dfvalidacion[self.dfvalidacion.columns[50]] = self.data[self.data.columns[50]].apply(
            lambda x: metodos.validarValorRespuesta(x, valores.dict_Tamizaje_TB))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[50],
                                 constansts.ERROR_VALORNOPERMITIDO, self.tipo)

    def validar_Antecedente_linfoma(self):
        self.dfvalidacion[self.dfvalidacion.columns[51]] = self.data[self.data.columns[51]].apply(
            lambda x: metodos.validarValorRespuesta(x, valores.dict_Antecedente_linfoma))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[51],
                                 constansts.ERROR_VALORNOPERMITIDO, self.tipo)

    def validar_Azatioprina(self):
        self.dfvalidacion[self.dfvalidacion.columns[52]] = self.data[self.data.columns[52]].apply(
            lambda x: metodos.validarValorRespuesta(x, valores.dict_Azatioprina))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[52],
                                 constansts.ERROR_VALORNOPERMITIDO, self.tipo)

    def validar_Ciclosporin(self):
        self.dfvalidacion[self.dfvalidacion.columns[53]] = self.data[self.data.columns[53]].apply(
            lambda x: metodos.validarValorRespuesta(x, valores.dict_Ciclosporina))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[53],
                                 constansts.ERROR_VALORNOPERMITIDO, self.tipo)

    def validar_Ciclofosfamida(self):
        self.dfvalidacion[self.dfvalidacion.columns[54]] = self.data[self.data.columns[54]].apply(
            lambda x: metodos.validarValorRespuesta(x, valores.dict_Ciclofosfamida))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[54],
                                 constansts.ERROR_VALORNOPERMITIDO, self.tipo)

    def validar_Cloroquina(self):
        self.dfvalidacion[self.dfvalidacion.columns[55]] = self.data[self.data.columns[55]].apply(
            lambda x: metodos.validarValorRespuesta(x, valores.dict_Cloroquina))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[55],
                                 constansts.ERROR_VALORNOPERMITIDO, self.tipo)

    def validar_D_penicilaimina(self):
        self.dfvalidacion[self.dfvalidacion.columns[56]] = self.data[self.data.columns[56]].apply(
            lambda x: metodos.validarValorRespuesta(x, valores.dict_D_penicilaimina))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[56],
                                 constansts.ERROR_VALORNOPERMITIDO, self.tipo)

    def validar_Etanercept(self):
        self.dfvalidacion[self.dfvalidacion.columns[57]] = self.data[self.data.columns[57]].apply(
            lambda x: metodos.validarValorRespuesta(x, valores.dict_Etanercept))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[57],
                                 constansts.ERROR_VALORNOPERMITIDO, self.tipo)

    def validar_Leflunomida(self):
        self.dfvalidacion[self.dfvalidacion.columns[58]] = self.data[self.data.columns[58]].apply(
            lambda x: metodos.validarValorRespuesta(x, valores.dict_Leflunomida))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[58],
                                 constansts.ERROR_VALORNOPERMITIDO, self.tipo)

    def validar_Metotrexate(self):
        self.dfvalidacion[self.dfvalidacion.columns[59]] = self.data[self.data.columns[59]].apply(
            lambda x: metodos.validarValorRespuesta(x, valores.dict_Metotrexate))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[59],
                                 constansts.ERROR_VALORNOPERMITIDO, self.tipo)

    def validar_Rituximab(self):
        self.dfvalidacion[self.dfvalidacion.columns[60]] = self.data[self.data.columns[60]].apply(
            lambda x: metodos.validarValorRespuesta(x, valores.dict_Rituximab))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[60],
                                 constansts.ERROR_VALORNOPERMITIDO, self.tipo)

    def validar_Sulfasalazina(self):
        self.dfvalidacion[self.dfvalidacion.columns[61]] = self.data[self.data.columns[61]].apply(
            lambda x: metodos.validarValorRespuesta(x, valores.dict_Sulfasalazina))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[61],
                                 constansts.ERROR_VALORNOPERMITIDO, self.tipo)

    def validar_Abatacept(self):
        self.dfvalidacion[self.dfvalidacion.columns[62]] = self.data[self.data.columns[62]].apply(
            lambda x: metodos.validarValorRespuesta(x, valores.dict_Abatacept))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[62],
                                 constansts.ERROR_VALORNOPERMITIDO, self.tipo)

    def validar_Adalimuma(self):
        self.dfvalidacion[self.dfvalidacion.columns[63]] = self.data[self.data.columns[63]].apply(
            lambda x: metodos.validarValorRespuesta(x, valores.dict_Adalimumab))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[63],
                                 constansts.ERROR_VALORNOPERMITIDO, self.tipo)

    def validar_Certolizumab(self):
        self.dfvalidacion[self.dfvalidacion.columns[64]] = self.data[self.data.columns[64]].apply(
            lambda x: metodos.validarValorRespuesta(x, valores.dict_Certolizumab))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[64],
                                 constansts.ERROR_VALORNOPERMITIDO, self.tipo)

    def validar_Golimumab(self):
        self.dfvalidacion[self.dfvalidacion.columns[65]] = self.data[self.data.columns[65]].apply(
            lambda x: metodos.validarValorRespuesta(x, valores.dict_Golimumab))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[65],
                                 constansts.ERROR_VALORNOPERMITIDO, self.tipo)

    def validar_Hidroxicloroquina(self):
        self.dfvalidacion[self.dfvalidacion.columns[66]] = self.data[self.data.columns[66]].apply(
            lambda x: metodos.validarValorRespuesta(x, valores.dict_Hidroxicloroquina))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[66],
                                 constansts.ERROR_VALORNOPERMITIDO, self.tipo)

    def validar_Infliximab(self):
        self.dfvalidacion[self.dfvalidacion.columns[67]] = self.data[self.data.columns[67]].apply(
            lambda x: metodos.validarValorRespuesta(x, valores.dict_Infliximab))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[67],
                                 constansts.ERROR_VALORNOPERMITIDO, self.tipo)

    def validar_Sales_oro(self):
        self.dfvalidacion[self.dfvalidacion.columns[68]] = self.data[self.data.columns[68]].apply(
            lambda x: metodos.validarValorRespuesta(x, valores.dict_Sales_oro))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[68],
                                 constansts.ERROR_VALORNOPERMITIDO, self.tipo)

    def validar_Tocilizumab(self):
        self.dfvalidacion[self.dfvalidacion.columns[69]] = self.data[self.data.columns[69]].apply(
            lambda x: metodos.validarValorRespuesta(x, valores.dict_Tocilizumab))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[69],
                                 constansts.ERROR_VALORNOPERMITIDO, self.tipo)

    def validar_Tofacitinib(self):
        self.dfvalidacion[self.dfvalidacion.columns[70]] = self.data[self.data.columns[70]].apply(
            lambda x: metodos.validarValorRespuesta(x, valores.dict_Tofacitinib))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[70],
                                 constansts.ERROR_VALORNOPERMITIDO, self.tipo)

    def validar_Anakinrab(self):
        self.dfvalidacion[self.dfvalidacion.columns[71]] = self.data[self.data.columns[71]].apply(
            lambda x: metodos.validarValorRespuesta(x, valores.dict_Anakinra))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[71],
                                 constansts.ERROR_VALORNOPERMITIDO, self.tipo)

    def validar_MedicamentoNoPos1(self):
        self.dfvalidacion[self.dfvalidacion.columns[72]] = self.data[self.data.columns[72]].apply(
            lambda x: metodos.validarValorRespuesta(x, valores.dict_medicamentosNoPos))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[72],
                                 constansts.ERROR_VALORNOPERMITIDO, self.tipo)


    def validar_MedicamentoNoPos2(self):
        self.dfvalidacion[self.dfvalidacion.columns[73]] = self.data[self.data.columns[73]].apply(
            lambda x: metodos.validarValorRespuesta(x, valores.dict_medicamentosNoPos))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[73],
                                 constansts.ERROR_VALORNOPERMITIDO, self.tipo)

    def validar_MedicamentoNoPos3(self):
        self.dfvalidacion[self.dfvalidacion.columns[74]] = self.data[self.data.columns[74]].apply(
            lambda x: metodos.validarValorRespuesta(x, valores.dict_medicamentosNoPos))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[74],
                                 constansts.ERROR_VALORNOPERMITIDO, self.tipo)


    def validar_MedicamentoNoPos4(self):
        self.dfvalidacion[self.dfvalidacion.columns[75]] = self.data[self.data.columns[75]].apply(
            lambda x: metodos.validarValorRespuesta(x, valores.dict_medicamentosNoPos))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[75],
                                 constansts.ERROR_VALORNOPERMITIDO, self.tipo)


    def validar_pesoUltimo(self):
        self.dfvalidacion[self.dfvalidacion.columns[76]] = self.data[self.data.columns[76]].apply(
            lambda x: metodos.validarRango(3, 300, x))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[76],
                                 constansts.ERROR_VALORNOPERMITIDO, self.tipo)

    def validar_Radiografia_manos_Ultima(self):
        self.dfvalidacion[self.dfvalidacion.columns[77]] = self.data[self.data.columns[77]].apply(
            lambda x: metodos.validarValorRespuesta(x, valores.dict_Radiografia_manos_Ultima))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[77],
                                 constansts.ERROR_VALORNOPERMITIDO, self.tipo)

    def validar_Radiografia_pies_Ultima(self):
        self.dfvalidacion[self.dfvalidacion.columns[78]] = self.data[self.data.columns[78]].apply(
            lambda x: metodos.validarValorRespuesta(x, valores.dict_Radiografia_pies_Ultima))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[78],
                                 constansts.ERROR_VALORNOPERMITIDO, self.tipo)

    def validar_PCRUltimo(self):
        self.dfvalidacion[self.dfvalidacion.columns[79]] = self.data[self.data.columns[79]].apply(
            lambda x: metodos.validarRangoSinDato(0, 250, 300, x))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[79],
                                 constansts.ERROR_VALORNOPERMITIDO, self.tipo)

    def validar_VSGUltimo(self):
        self.dfvalidacion[self.dfvalidacion.columns[80]] = self.data[self.data.columns[80]].apply(
            lambda x: metodos.validarRangoSinDato(0, 250, 300, x))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[80],
                                 constansts.ERROR_VALORNOPERMITIDO, self.tipo)

    def validar_HemoglobinaUltimo(self):
        self.dfvalidacion[self.dfvalidacion.columns[81]] = self.data[self.data.columns[81]].apply(
            lambda x: metodos.validarRangoSinDato(0, 50, 300, x))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[81],
                                 constansts.ERROR_VALORNOPERMITIDO, self.tipo)

    def validar_LecocitosUltimo(self):
        self.dfvalidacion[self.dfvalidacion.columns[82]] = self.data[self.data.columns[82]].apply(
            lambda x: metodos.validarRangoSinDato(0, 20000, 22222, x))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[82],
                                 constansts.ERROR_VALORNOPERMITIDO, self.tipo)

    def validar_CreatininaUltimo(self):
        self.dfvalidacion[self.dfvalidacion.columns[83]] = self.data[self.data.columns[83]].apply(
            lambda x: metodos.validarRangoSinDato(0, 20, 300, x))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[83],
                                 constansts.ERROR_VALORNOPERMITIDO, self.tipo)

    def validar_TFGULtimo(self):
        self.dfvalidacion[self.dfvalidacion.columns[84]] = self.data[self.data.columns[84]].apply(
            lambda x: metodos.validarRangoSinDato(0, 250, 300, x))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[84],
                                 constansts.ERROR_VALORNOPERMITIDO, self.tipo)

    def validar_Parcial_Orina_ultimo(self):
        self.dfvalidacion[self.dfvalidacion.columns[85]] = self.data[self.data.columns[85]].apply(
            lambda x: metodos.validarValorRespuesta(x, valores.dict_Parcial_Orina_ultimo))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[85],
                                 constansts.ERROR_VALORNOPERMITIDO, self.tipo)

    def validar_ALT_ultimo(self):
        self.dfvalidacion[self.dfvalidacion.columns[86]] = self.data[self.data.columns[86]].apply(
            lambda x: metodos.validarValorRespuesta(x, valores.dict_ALT_ultimo))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[86],
                                 constansts.ERROR_VALORNOPERMITIDO, self.tipo)

    def validar_HTA_Actual(self):
        self.dfvalidacion[self.dfvalidacion.columns[87]] = self.data[self.data.columns[87]].apply(
            lambda x: metodos.validarValorRespuesta(x, valores.dict_HTA_Actual))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[87],
                                 constansts.ERROR_VALORNOPERMITIDO, self.tipo)

    def validar_DM_Actual(self):
        self.dfvalidacion[self.dfvalidacion.columns[88]] = self.data[self.data.columns[88]].apply(
            lambda x: metodos.validarValorRespuesta(x, valores.dict_DM_Actual))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[88],
                                 constansts.ERROR_VALORNOPERMITIDO, self.tipo)

    def validar_ECV_Actual(self):
        self.dfvalidacion[self.dfvalidacion.columns[89]] = self.data[self.data.columns[89]].apply(
            lambda x: metodos.validarValorRespuesta(x, valores.dict_ECV_Actual))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[89],
                                 constansts.ERROR_VALORNOPERMITIDO, self.tipo)

    def validar_ERC_Actual(self):
        self.dfvalidacion[self.dfvalidacion.columns[90]] = self.data[self.data.columns[90]].apply(
            lambda x: metodos.validarValorRespuesta(x, valores.dict_ERC_Actual))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[90],
                                 constansts.ERROR_VALORNOPERMITIDO, self.tipo)

    def validar_Osteoporosis_Actual(self):
        self.dfvalidacion[self.dfvalidacion.columns[91]] = self.data[self.data.columns[91]].apply(
            lambda x: metodos.validarValorRespuesta(x, valores.dict_Osteoporosis_Actual))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[91],
                                 constansts.ERROR_VALORNOPERMITIDO, self.tipo)

    def validar_Sindrome_Sjogren_Actual(self):
        self.dfvalidacion[self.dfvalidacion.columns[92]] = self.data[self.data.columns[92]].apply(
            lambda x: metodos.validarValorRespuesta(x, valores.dict_Sindrome_Sjogren_Actual))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[92],
                                 constansts.ERROR_VALORNOPERMITIDO, self.tipo)

    def validar_FechaUltimaDAS(self):
        self.dfvalidacion[self.dfvalidacion.columns[93]] = self.data[self.data.columns[93]].apply(
            lambda x: metodos.validarFechaSinDato(x))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[93],
                                 constansts.ERROR_FECHA_SINDATO, self.tipo)

    def validar_Profesional_DAS_Ultimo(self):
        self.dfvalidacion[self.dfvalidacion.columns[94]] = self.data[self.data.columns[94]].apply(
            lambda x: metodos.validarValorRespuesta(x, valores.dict_Profesional_DAS_Ultimo))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[94],
                                 constansts.ERROR_VALORNOPERMITIDO, self.tipo)

    def validar_TFGULtimodos(self):
        self.dfvalidacion[self.dfvalidacion.columns[95]] = self.data[self.data.columns[95]].apply(
            lambda x: metodos.validarRangoSinDato(0, 10, 300, x))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[95],
                                 constansts.ERROR_VALORNOPERMITIDO, self.tipo)

    def validar_Estado_AR(self):
        self.dfvalidacion[self.dfvalidacion.columns[96]] = self.data[self.data.columns[96]].apply(
            lambda x: metodos.validarValorRespuesta(x, valores.dict_Estado_AR))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[96],
                                 constansts.ERROR_VALORNOPERMITIDO, self.tipo)

    def validar_FechaUltimaHAQ(self):
        self.dfvalidacion[self.dfvalidacion.columns[97]] = self.data[self.data.columns[97]].apply(
            lambda x: metodos.validarFechaSinDato(x))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[97],
                                 constansts.ERROR_FECHA_SINDATO, self.tipo)

    def validar_HAQUltimo(self):
        self.dfvalidacion[self.dfvalidacion.columns[98]] = self.data[self.data.columns[98]].apply(
            lambda x: metodos.validarRangoSinDato(0, 3, 300, x))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[98],
                                 constansts.ERROR_VALORNOPERMITIDO, self.tipo)

    def validar_Analgesicos_No_Opioides_Dos(self):
        self.dfvalidacion[self.dfvalidacion.columns[99]] = self.data[self.data.columns[99]].apply(
            lambda x: metodos.validarValorRespuesta(x, valores.dict_Analgesicos_No_Opioides_Dos))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[99],
                                 constansts.ERROR_VALORNOPERMITIDO, self.tipo)

    def validar_Analgesicos_Opioides_Dos(self):
        self.dfvalidacion[self.dfvalidacion.columns[100]] = self.data[self.data.columns[100]].apply(
            lambda x: metodos.validarValorRespuesta(x, valores.dict_Analgesicos_Opioides_Dos))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[100],
                                 constansts.ERROR_VALORNOPERMITIDO, self.tipo)

    def validar_AINES_Do(self):
        self.dfvalidacion[self.dfvalidacion.columns[101]] = self.data[self.data.columns[101]].apply(
            lambda x: metodos.validarValorRespuesta(x, valores.dict_AINES_Dos))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[101],
                                 constansts.ERROR_VALORNOPERMITIDO, self.tipo)

    def validar_Corticoides_Dos(self):
        self.dfvalidacion[self.dfvalidacion.columns[102]] = self.data[self.data.columns[102]].apply(
            lambda x: metodos.validarValorRespuesta(x, valores.dict_Corticoides_Dos))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[102],
                                 constansts.ERROR_VALORNOPERMITIDO, self.tipo)

    def validar_glucocorticoides_dosis(self):
        self.dfvalidacion[self.dfvalidacion.columns[103]] = self.data[self.data.columns[102]].apply(
            lambda x: metodos.validarRango(0, 12, x))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[103],
                                 constansts.ERROR_VALORNOPERMITIDO, self.tipo)

    def validar_Calcio(self):
        self.dfvalidacion[self.dfvalidacion.columns[104]] = self.data[self.data.columns[104]].apply(
            lambda x: metodos.validarValorRespuesta(x, valores.dict_Calcio))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[104],
                                 constansts.ERROR_VALORNOPERMITIDO, self.tipo)

    def validar_Vitamina_D(self):
        self.dfvalidacion[self.dfvalidacion.columns[105]] = self.data[self.data.columns[105]].apply(
            lambda x: metodos.validarValorRespuesta(x, valores.dict_Vitamina_D))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[105],
                                 constansts.ERROR_VALORNOPERMITIDO, self.tipo)

    def validar_FechaINicioDMARD(self):
        self.dfvalidacion[self.dfvalidacion.columns[106]] = self.data[self.data.columns[106]].apply(
            lambda x: metodos.validarFechaSinDato(x))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[106],
                                 constansts.ERROR_FECHA_SINDATO, self.tipo)

    def validar_Azatioprina_Dos(self):
        self.dfvalidacion[self.dfvalidacion.columns[107]] = self.data[self.data.columns[107]].apply(
            lambda x: metodos.validarValorRespuesta(x, valores.dict_Azatioprina_Dos))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[107],
                                 constansts.ERROR_VALORNOPERMITIDO, self.tipo)

    def validar_Ciclosporina_Dos(self):
        self.dfvalidacion[self.dfvalidacion.columns[108]] = self.data[self.data.columns[108]].apply(
            lambda x: metodos.validarValorRespuesta(x, valores.dict_Ciclosporina_Dos))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[108],
                                 constansts.ERROR_VALORNOPERMITIDO, self.tipo)

    def validar_Ciclofosfamida_Dos(self):
        self.dfvalidacion[self.dfvalidacion.columns[109]] = self.data[self.data.columns[109]].apply(
            lambda x: metodos.validarValorRespuesta(x, valores.dict_Ciclofosfamida_Dos))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[109],
                                 constansts.ERROR_VALORNOPERMITIDO, self.tipo)

    def validar_Cloroquina_DOS(self):
        self.dfvalidacion[self.dfvalidacion.columns[110]] = self.data[self.data.columns[110]].apply(
            lambda x: metodos.validarValorRespuesta(x, valores.dict_Cloroquina_DOS))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[110],
                                 constansts.ERROR_VALORNOPERMITIDO, self.tipo)

    def validar_D_penicilaimina_Dos(self):
        self.dfvalidacion[self.dfvalidacion.columns[111]] = self.data[self.data.columns[111]].apply(
            lambda x: metodos.validarValorRespuesta(x, valores.dict_D_penicilaimina_Dos))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[111],
                                 constansts.ERROR_VALORNOPERMITIDO, self.tipo)

    def validar_Etanercept_Dos(self):
        self.dfvalidacion[self.dfvalidacion.columns[112]] = self.data[self.data.columns[112]].apply(
            lambda x: metodos.validarValorRespuesta(x, valores.dict_Etanercept_Dos))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[112],
                                 constansts.ERROR_VALORNOPERMITIDO, self.tipo)

    def validar_Leflunomida_Dos(self):
        self.dfvalidacion[self.dfvalidacion.columns[113]] = self.data[self.data.columns[113]].apply(
            lambda x: metodos.validarValorRespuesta(x, valores.dict_Leflunomida_Dos))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[113],
                                 constansts.ERROR_VALORNOPERMITIDO, self.tipo)

    def validar_Metotrexate_Dos(self):
        self.dfvalidacion[self.dfvalidacion.columns[114]] = self.data[self.data.columns[114]].apply(
            lambda x: metodos.validarValorRespuesta(x, valores.dict_Metotrexate_Dos))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[114],
                                 constansts.ERROR_VALORNOPERMITIDO, self.tipo)

    def validar_Rituximab_Dos(self):
        self.dfvalidacion[self.dfvalidacion.columns[115]] = self.data[self.data.columns[115]].apply(
            lambda x: metodos.validarValorRespuesta(x, valores.dict_Rituximab_Dos))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[115],
                                 constansts.ERROR_VALORNOPERMITIDO, self.tipo)

    def validar_Sulfasalazina_Dos(self):
        self.dfvalidacion[self.dfvalidacion.columns[116]] = self.data[self.data.columns[116]].apply(
            lambda x: metodos.validarValorRespuesta(x, valores.dict_Sulfasalazina_Dos))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[116],
                                 constansts.ERROR_VALORNOPERMITIDO, self.tipo)

    def validar_Abatacept_Dos(self):
        self.dfvalidacion[self.dfvalidacion.columns[117]] = self.data[self.data.columns[117]].apply(
            lambda x: metodos.validarValorRespuesta(x, valores.dict_Abatacept_Dos))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[117],
                                 constansts.ERROR_VALORNOPERMITIDO, self.tipo)

    def validar_Adalimumab_Dos(self):
        self.dfvalidacion[self.dfvalidacion.columns[118]] = self.data[self.data.columns[118]].apply(
            lambda x: metodos.validarValorRespuesta(x, valores.dict_Adalimumab_Dos))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[118],
                                 constansts.ERROR_VALORNOPERMITIDO, self.tipo)

    def validar_Certolizumab_Dos(self):
        self.dfvalidacion[self.dfvalidacion.columns[119]] = self.data[self.data.columns[119]].apply(
            lambda x: metodos.validarValorRespuesta(x, valores.dict_Certolizumab_Dos))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[119],
                                 constansts.ERROR_VALORNOPERMITIDO, self.tipo)

    def validar_Golimumab_Dos(self):
        self.dfvalidacion[self.dfvalidacion.columns[120]] = self.data[self.data.columns[120]].apply(
            lambda x: metodos.validarValorRespuesta(x, valores.dict_Golimumab_Dos))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[120],
                                 constansts.ERROR_VALORNOPERMITIDO, self.tipo)

    def validar_Hidroxicloroquina_Dos(self):
        self.dfvalidacion[self.dfvalidacion.columns[121]] = self.data[self.data.columns[121]].apply(
            lambda x: metodos.validarValorRespuesta(x, valores.dict_Hidroxicloroquina_Dos))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[121],
                                 constansts.ERROR_VALORNOPERMITIDO, self.tipo)

    def validar_Infliximab_Dos(self):
        self.dfvalidacion[self.dfvalidacion.columns[122]] = self.data[self.data.columns[122]].apply(
            lambda x: metodos.validarValorRespuesta(x, valores.dict_Infliximab_Dos))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[122],
                                 constansts.ERROR_VALORNOPERMITIDO, self.tipo)

    def validar_Sales_oro_dos(self):
        self.dfvalidacion[self.dfvalidacion.columns[123]] = self.data[self.data.columns[123]].apply(
            lambda x: metodos.validarValorRespuesta(x, valores.dict_Sales_oro_dos))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[123],
                                 constansts.ERROR_VALORNOPERMITIDO, self.tipo)

    def validar_Tocilizumab_Dos(self):
        self.dfvalidacion[self.dfvalidacion.columns[124]] = self.data[self.data.columns[124]].apply(
            lambda x: metodos.validarValorRespuesta(x, valores.dict_Tocilizumab_Dos))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[124],
                                 constansts.ERROR_VALORNOPERMITIDO, self.tipo)

    def validar_Tofacitinib_Dos(self):
        self.dfvalidacion[self.dfvalidacion.columns[125]] = self.data[self.data.columns[125]].apply(
            lambda x: metodos.validarValorRespuesta(x, valores.dict_Tofacitinib_Dos))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[125],
                                 constansts.ERROR_VALORNOPERMITIDO, self.tipo)

    def validar_Anakinra_Dos(self):
        self.dfvalidacion[self.dfvalidacion.columns[126]] = self.data[self.data.columns[126]].apply(
            lambda x: metodos.validarValorRespuesta(x, valores.dict_Anakinra_Dos))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[126],
                                 constansts.ERROR_VALORNOPERMITIDO, self.tipo)

    def validar_OtroMedicamentoNoPos1(self):
        self.dfvalidacion[self.dfvalidacion.columns[127]] = self.data[self.data.columns[127]].apply(
            lambda x: metodos.validarValorRespuesta(x, valores.dict_medicamentosNoPos))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[127],
                                 constansts.ERROR_VALORNOPERMITIDO, self.tipo)


    def validar_OtroMedicamentoNoPos2(self):
        self.dfvalidacion[self.dfvalidacion.columns[128]] = self.data[self.data.columns[128]].apply(
            lambda x: metodos.validarValorRespuesta(x, valores.dict_medicamentosNoPos))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[128],
                                 constansts.ERROR_VALORNOPERMITIDO, self.tipo)


    def validar_OtroMedicamentoNoPos3(self):
        self.dfvalidacion[self.dfvalidacion.columns[129]] = self.data[self.data.columns[129]].apply(
            lambda x: metodos.validarValorRespuesta(x, valores.dict_medicamentosNoPos))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[129],
                                 constansts.ERROR_VALORNOPERMITIDO, self.tipo)


    def validar_OtroMedicamentoNoPos4(self):
        self.dfvalidacion[self.dfvalidacion.columns[130]] = self.data[self.data.columns[130]].apply(
            lambda x: metodos.validarValorRespuesta(x, valores.dict_medicamentosNoPos))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[130],
                                 constansts.ERROR_VALORNOPERMITIDO, self.tipo)


    def validar_numeroReumatologo(self):
        self.dfvalidacion[self.dfvalidacion.columns[131]] = self.data[self.data.columns[131]].apply(
            lambda x: metodos.validarRango(0, 12, x))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[131],
                                 constansts.ERROR_VALORNOPERMITIDO, self.tipo)

    def validar_numeroInternistaAr(self):
        self.dfvalidacion[self.dfvalidacion.columns[132]] = self.data[self.data.columns[132]].apply(
            lambda x: metodos.validarRango(0, 12, x))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[132],
                                 constansts.ERROR_VALORNOPERMITIDO, self.tipo)

    def validar_numeroConsultasMedicoAR(self):
        self.dfvalidacion[self.dfvalidacion.columns[133]] = self.data[self.data.columns[133]].apply(
            lambda x: metodos.validarRango(0, 12, x))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[133],
                                 constansts.ERROR_VALORNOPERMITIDO, self.tipo)

    def validar_Reemplazo_articular_1_por_AR(self):
        self.dfvalidacion[self.dfvalidacion.columns[134]] = self.data[self.data.columns[134]].apply(
            lambda x: metodos.validarValorRespuesta(x, valores.dict_Reemplazo_articular_1_por_AR))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[134],
                                 constansts.ERROR_VALORNOPERMITIDO, self.tipo)

    def validar_Reemplazo_articular_2_por_AR(self):
        self.dfvalidacion[self.dfvalidacion.columns[135]] = self.data[self.data.columns[135]].apply(
            lambda x: metodos.validarValorRespuesta(x, valores.dict_Reemplazo_articular_2_por_AR))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[135],
                                 constansts.ERROR_VALORNOPERMITIDO, self.tipo)

    def validar_Reemplazo_articular_3_por_AR(self):
        self.dfvalidacion[self.dfvalidacion.columns[136]] = self.data[self.data.columns[136]].apply(
            lambda x: metodos.validarValorRespuesta(x, valores.dict_Reemplazo_articular_3_por_AR))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[136],
                                 constansts.ERROR_VALORNOPERMITIDO, self.tipo)

    def validar_Reemplazo_articular_4_por_AR(self):
        self.dfvalidacion[self.dfvalidacion.columns[137]] = self.data[self.data.columns[137]].apply(
            lambda x: metodos.validarValorRespuesta(x, valores.dict_Reemplazo_articular_4_por_AR))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[137],
                                 constansts.ERROR_VALORNOPERMITIDO, self.tipo)

    def validar_numeroHospitalizacionesAR(self):
        self.dfvalidacion[self.dfvalidacion.columns[138]] = self.data[self.data.columns[138]].apply(
            lambda x: metodos.validarRango(0, 12, x))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[138],
                                 constansts.ERROR_VALORNOPERMITIDO, self.tipo)

    def validar_CodigoHailitacion(self):
        print(self.data[self.data.columns[139]])
        self.dfvalidacion[self.dfvalidacion.columns[139]] = self.data[self.data.columns[139]].apply(
            lambda x: metodos.validarValorRespuesta(str(x), valores.dict_IPS))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[139],
                                 constansts.ERROR_VALORNOPERMITIDO, self.tipo)

    def validar_CodigoMunicipioips(self):
        self.dfvalidacion[self.dfvalidacion.columns[140]] = self.data[self.data.columns[140]].apply(
            lambda x: metodos.validarValorRespuesta(str(x), valores.dict_mucnicipio_dos))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[140],
                                 constansts.ERROR_VALORNOPERMITIDO, self.tipo)

    def validar_fechaIngresoIPS(self):
        self.dfvalidacion[self.dfvalidacion.columns[141]] = self.data[self.data.columns[141]].apply(
            lambda x: metodos.validarFechaSinDato(x))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[141],
                                 constansts.ERROR_FECHA_SINDATO, self.tipo)

    def validar_Atencion_Clinica_AR(self):
        self.dfvalidacion[self.dfvalidacion.columns[142]] = self.data[self.data.columns[142]].apply(
            lambda x: metodos.validarValorRespuesta(x, valores.dict_Atencion_Clinica_AR))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[142],
                                 constansts.ERROR_VALORNOPERMITIDO, self.tipo)

    def validar_Novedad_Paciente_Reporte_Anterior(self):
        self.dfvalidacion[self.dfvalidacion.columns[143]] = self.data[self.data.columns[143]].apply(
            lambda x: metodos.validarValorRespuesta(x, valores.dict_Novedad_Paciente_Reporte_Anterior))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[143],
                                 constansts.ERROR_VALORNOPERMITIDO, self.tipo)

    def validar_FechaDesafiliacion(self):
        self.dfvalidacion[self.dfvalidacion.columns[144]] = self.data[self.data.columns[144]].apply(
            lambda x: metodos.validarFechaPorDefecto(x, '1799-01-01'))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[144],
                                 constansts.ERROR_FECHA_SINDATO, self.tipo)

    def validar_TrasladoPaciente(self):
        self.dfvalidacion[self.dfvalidacion.columns[145]] = self.data[self.data.columns[145]].apply(
            lambda x: metodos.validarValorRespuesta(str(x), valores.dict_trasladoPaciente))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[145],
                                 constansts.ERROR_VALORNOPERMITIDO, self.tipo)

    def validar_FechaMuerte(self):
        self.dfvalidacion[self.dfvalidacion.columns[146]] = self.data[self.data.columns[146]].apply(
            lambda x: metodos.validarFechaPorDefecto(x, '1799-01-01'))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[19],
                                 constansts.ERROR_FECHA_SINDATO, self.tipo)

    def validar_CausaMuerte(self):
        self.dfvalidacion[self.dfvalidacion.columns[147]] = self.data[self.data.columns[147]].apply(
            lambda x: metodos.validarValorRespuesta(x, valores.dict_causaMuerte))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[147],
                                 constansts.ERROR_VALORNOPERMITIDO,self.tipo)

    def validar_CostoDMARDPOS(self):
        self.dfvalidacion[self.dfvalidacion.columns[148]] = self.data[self.data.columns[148]].apply(
            lambda x: metodos.validar_SindecimalesSinComas(x))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[148],
                                 constansts.ERROR_SOLONUMEROS, self.tipo)

    def validar_CostoDMARDNoPOS(self):
        self.dfvalidacion[self.dfvalidacion.columns[149]] = self.data[self.data.columns[149]].apply(
            lambda x: metodos.validar_SindecimalesSinComas(x))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[149],
                                 constansts.ERROR_SOLONUMEROS, self.tipo)

    def validar_costoAnulaAR(self):
        self.dfvalidacion[self.dfvalidacion.columns[150]] = self.data[self.data.columns[150]].apply(
            lambda x: metodos.validar_SindecimalesSinComas(x))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[150],
                                 constansts.ERROR_SOLONUMEROS, self.tipo)

    def validad_costoAnualIncapacidad(self):
        self.dfvalidacion[self.dfvalidacion.columns[151]] = self.data[self.data.columns[151]].apply(
            lambda x: metodos.validar_SindecimalesSinComas(x))

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[150],
                                 constansts.ERROR_SOLONUMEROS, self.tipo)


    def validar_tipoidentificacionregimen(self):
        self.dfvalidacion[self.dfvalidacion.columns[7]] = self.data.apply(
            metodos.validartipoidentitifacion,axis=1)

        self.crearInformeErrores(self.dfvalidacion, self.data, self._columnas[7],
                                 constansts.ERROR_REGIMEN_TIPOIDENTIFICACION, self.tipo)
